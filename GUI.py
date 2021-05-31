import pandas as pd
import numpy as np
from tkinter import filedialog
from tkinter import *
import os
from os.path import isfile, join
import matplotlib.pyplot as plt
from math import atan,pi
from openpyxl import load_workbook
from openpyxl.styles import Alignment



numberofpoints=numberofpoints2=numberofpoints3=0

class point:
    # Define an object Point that contains the point's coordinates and intensity
    def __init__(self,x,y,z,intensity):
        self.x=x
        self.y=y
        self.z=z
        self.intensity=intensity
        if (x != 0):
            self.new_int=((x*x+y*y+z*z)/(x*x))*intensity

def read_file(path):
    #function to read the data from file and save it into a list of points
    liste=[]
    f1=pd.read_csv(path,sep=' ',header = None,names=['X','Y','Z','Intensity'])

    i=0
    while(i<len(f1)):        
        x_value=float(f1.iloc[i]['X'])
        y_value=float(f1.iloc[i]['Y'])
        z_value=float(f1.iloc[i]['Z'])
        intensity_value=float(f1.iloc[i]['Intensity'])
        liste.append(point(x_value,y_value,z_value,intensity_value) )
        i=i+1
    
    return liste            

def browse_savefolder():
    # Allow user to select a directory and store it in global var
    # called folder_path
    global save_path
    save_path = filedialog.askdirectory()
    save_var.set(save_path)

def savingdirectory():
    select_saving_path = Label(root,text ="3. Select saving directory:")
    select_saving_path.place(x=10,y=120, anchor='w')
    global save_var
    save_var= StringVar()
    save_var.set('No Folder Selected yet')
    
    save_directory_browse = Button(text="Browse",bg='orange', command=lambda:browse_savefolder())
    save_directory_browse.place(x=155, y=120, anchor='w')
    lbl2 = Label(master=root,textvariable=save_var)
    lbl2.place(x=220, y=120, anchor='w')

def browsecallback(selection):
    
    global selected_file
    selected_file = selection


def browse_for_path():
    try:
        folder_path.set(filedialog.askdirectory())
        folder=folder_path.get()
        onlyfiles = [f for f in os.listdir(folder_path.get()) if isfile(join(folder_path.get(), f))]
        csvfiles = []
        for i in range (len(onlyfiles)):
            if ((onlyfiles[i][-3:])=='csv'):
                csvfiles.append(onlyfiles[i])
        file = StringVar()
        file.set('No File selected yet') # default value
        filemenu = OptionMenu(root, file, *csvfiles,command=browsecallback)  
        filemenu.place(x=270, y=83, anchor='se')
    except(RuntimeError, TypeError, NameError,ValueError) :
        lbl2 = Label(master=root,text='ERROR : NO CSV FILES IN CHOSEN DIRECTORY !!!!')
        lbl2.config(font=("Courier", 16),fg='red')
        lbl2.place(x=10, y=120, anchor='w')
        lbl2.after(2500, lbl2.place_forget)
        
def get_new_path():
    label=Label(root,text='something was selected')
    label.place(x=120,y=69, anchor='w')
def folderselection():

    select_folder_text = Label(root, text="1. Select Folder :")
    select_folder_text.place(x=100, y=30, anchor='se')    
    
    global folder_path
    folder_path = StringVar()
    folder_path.set('No Folder Selected yet')
    
    choose_button = Button(text="Browse",bg='orange', command=lambda:browse_for_path())
    choose_button.place(x=170, y=35, anchor='se')  
    
    folder_path_text = Label(root,textvariable=folder_path)
    folder_path_text.place(x=180, y=20, anchor='w')  

def fileselection():
    Select_File_Text = Label(root, text="2. Select The file :")
    Select_File_Text.place(x=106,y=80, anchor='se')


def plot_dist(file_path,selected_file):

    path=file_path +'\\'+ selected_file
    points=read_file(path)
    xvalues=[]
    for i in range(len(points)):
        xvalues.append(points[i].x)
    x = np.asarray(xvalues, dtype=np.float32)    
    plt.hist(x)
    plt.gca().set_title('Distance Histogramm',loc='left')
    plt.axvline(x.mean(), color='k', linestyle='dashed', linewidth=1)
    min_ylim, max_ylim = plt.ylim()
    plt.text(x.mean()*1, max_ylim*1.01, 'Mean: {:.2f}'.format(x.mean()))
    plt.show()  

def plot_dist_button():
    button3 = Button(text="Plot Distance Histogramm",command=lambda:plot_dist(folder_path.get(),selected_file))
    button3.place(relx=0.2, y=180, anchor='se')    

def update_me(v_gotten,h_gotten,n_gotten,root):
    global target_v_orig
    global target_h_orig
    global number_of_frames


    target_v_orig=float(v_gotten)
    target_h_orig=float(h_gotten)
    number_of_frames=float(n_gotten)
    lbl = Label(root, text="Values Updated !!")
    lbl.place(x=657, y=135, anchor='w')
    lbl.after(1500, lbl.place_forget)
    
def berech(File,mypath):
 
    
    path= mypath +'/'+ File
    points=read_file(str(path))
    xvalues=[]
    yvalues=[]
    zvalues=[]
    inttvalues=[]
    for i in range(len(points)):
        xvalues.append(points[i].x)
        yvalues.append(points[i].y)
        zvalues.append(points[i].z)
        inttvalues.append(points[i].intensity)
        
    x = np.asarray(xvalues, dtype=np.float32)
    y = np.asarray(yvalues, dtype=np.float32)
    z = np.asarray(zvalues, dtype=np.float32)
    intt = np.asarray(inttvalues, dtype=np.float32)
    numberofpoints=len(x)
    target_h = np.max(y) - np.min(y)                                
    alpha = (atan(target_h / np.mean(x)))*(180/pi) 
    pxl_soll_h = np.ceil(alpha / h_res)   +1                        
    
    target_v = np.max(z) - np.min(z)
    alpha2 = (atan(target_v / np.mean(x)))*(180/pi) 
    pxl_soll_v = np.ceil(alpha2 / v_res)  +1                          
    pxl_soll = pxl_soll_h * pxl_soll_v                                
    
    detect_prob = (numberofpoints/number_of_frames) / pxl_soll 
    
    calcul = {'Number of points in File': [numberofpoints],'Farthest detected Point':[np.max(x)],
           'Nearest detected Point': [np.min(x)],'Mean of Distance values':[np.mean(x)],
           'Distance Median': [np.median(x)],'Standard Deviation of Distance values':[np.std(x)],
           'Highest Intensity value': [np.max(intt)],'Lowest Intensity value':[np.min(intt)],
           'Mean of Intensity': [np.mean(intt)],'Intensity Median':[np.median(intt)],
           'Standard Deviation of Intensity': [np.std(intt)],'Calculated number of h-Pixels':[pxl_soll_h],
           'Calculated number of v-Pixels': [pxl_soll_v],'Calculated number of Pixels per Frame':[pxl_soll],
           'Number of Pixels per Frame': [np.ceil(numberofpoints/number_of_frames)],
           'Detection Propability (%)': [np.round(detect_prob * 100)] }
    df = pd.DataFrame(calcul)

    return df

def compute_data_from_all_files(mypath,dir_name): 
    files = [f for f in os.listdir(mypath) if isfile(join(mypath, f))]

    List = []
    for i in range (len(files)):
        if ((files[i][-3:])=='csv'):
            List.append(files[i])

    for i in range (len(List)):
        current=List[i]
        try:
            Distance = int(current[:5])
        except(RuntimeError, TypeError, NameError,ValueError):
            Distance = 'NaN'
        try:
            reflectivity = int(current[6:8])
        except(RuntimeError, TypeError, NameError,ValueError):
            reflectivity ='NaN'
        try:
            winkelh = int(current[8:11])
        except(RuntimeError, TypeError, NameError,ValueError):
            winkelh= 'NaN'
        try:
            winkelv = int(current[11:14])
        except(RuntimeError, TypeError, NameError,ValueError):
            winkelv='NaN'
        cal = {'File name :':[current],'Test Distance in cm': [Distance],'Test Reflectivity':[reflectivity],
           'Horizontal Angle': [winkelh],'Vertical Angle':[winkelv]}
        dftest = pd.DataFrame(cal)
        df= berech(current,mypath)
        result = pd.concat([dftest,df], axis=1, join='outer', ignore_index=False, keys=None,
                   levels=None, names=None, verify_integrity=False, copy=True)
        if (i == 0):
            tabelle = result
        else:
            tabelle = pd.concat([tabelle,result], axis=0, join='outer', ignore_index=False, keys=None,
                                levels=None, names=None, verify_integrity=False, copy=True)

    tabelle.to_excel (r'{}/export_dataframe.xlsx'.format(dir_name), index = None, header=True)

    wb = load_workbook(filename = '{}/export_dataframe.xlsx'.format(dir_name))
    sheet = wb.active
    
        
    for c in range(ord('A'), ord('Z')+1):
        sheet.column_dimensions[chr(c)].width = 20    
        var=str(chr(c)+'1')
        sheet[var].alignment = Alignment(wrapText=True)

    wb.save('{}/export_dataframe.xlsx'.format(dir_name)) 
    os.system("start "+"{}/export_dataframe.xlsx".format(dir_name))  
       
def one_file(mypath,File):
    save_dir=filedialog.askdirectory()
    current=File
    try:
        Distance = int(current[:5])
    except(RuntimeError, TypeError, NameError,ValueError):
        Distance = 'NaN'
    try:
        reflectivity = int(current[6:8])
    except(RuntimeError, TypeError, NameError,ValueError):
        reflectivity ='NaN'
    try:
        winkelh = int(current[8:11])
    except(RuntimeError, TypeError, NameError,ValueError):
        winkelh= 'NaN'
    try:
        winkelv = int(current[11:14])
    except(RuntimeError, TypeError, NameError,ValueError):
        winkelv='NaN'
    cal = {'File name :':[current],'Test Distance in cm': [Distance],'Test Reflectivity':[reflectivity],
       'Horizontal Angle': [winkelh],'Vertical Angle':[winkelv]}
    dftest = pd.DataFrame(cal)
    df= berech(File,mypath)
    result = pd.concat([dftest,df], axis=1, join='outer', ignore_index=False, keys=None,
               levels=None, names=None, verify_integrity=False, copy=True)
    result.to_excel (r"{}/{}.xlsx".format(save_dir,File), index = None, header=True)
    wb = load_workbook(filename = "{}/{}.xlsx".format(save_dir,File))
    sheet = wb.active
    sheet.row_dimensions[1].height = 20    
    for c in range(ord('A'), ord('Z')+1):
        sheet.column_dimensions[chr(c)].width = 20    
        var=str(chr(c)+'1')
        sheet[var].alignment = Alignment(wrapText=True)   
    wb.save("{}/{}.xlsx".format(save_dir,File)) 
    os.system("start "+"{}/{}.xlsx".format(save_dir,File))


def entry_buttons_show():
    

    target_v_orig_var=StringVar()
    target_v_orig_var.set(80)
    target_v_orig_entry = Entry(root, textvariable=target_v_orig_var)
    target_v_orig_entry.place(x=670,y=30,anchor='w')
    lbl10 = Label(root, text="Target Height(cm)")
    lbl10.place(x=559, y=30, anchor='w')
    
    target_h_orig_var=StringVar()
    target_h_orig_var.set(50)
    target_h_orig_entry = Entry(root, textvariable=target_h_orig_var)
    target_h_orig_entry.place(x=670,y=50,anchor='w')
    lbl20 = Label(root, text="Target Width(cm)")
    lbl20.place(x=559, y=50, anchor='w')
    
    number_of_frames_var=StringVar()
    number_of_frames_var.set(10)
    number_of_frames_entry = Entry(root, textvariable=number_of_frames_var)
    number_of_frames_entry.place(x=670,y=70,anchor='w')
    lbl30 = Label(root, text='N of Frames per File')
    lbl30.place(x=559, y=70, anchor='w')
    



    target_v_orig=float(target_v_orig_var.get())
    target_h_orig=float(target_h_orig_var.get())
    number_of_frames=float(number_of_frames_var.get())

    test_button = Button(text='UPDATE' , command=lambda:update_me(target_v_orig_entry.get(),target_h_orig_entry.get(),number_of_frames_entry.get(),root))
    test_button.place(x=600,y=135,anchor='w')    
def plot_int(file_path,selected_file):

    path=file_path +'\\'+ selected_file
    points=read_file(path)
    intvalues=[]
    for i in range(len(points)):
        intvalues.append(points[i].intensity)
    x = np.asarray(intvalues, dtype=np.float32)    
    plt.hist(x)
    plt.gca().set_title('Distance Histogramm',loc='left')
    plt.axvline(x.mean(), color='k', linestyle='dashed', linewidth=1)
    min_ylim, max_ylim = plt.ylim()
    plt.text(x.mean()*1, max_ylim*1.01, 'Mean: {:.2f}'.format(x.mean()))
    plt.show()  
    
def plot_int_button():
    button4 = Button(text="Plot Intensity Histogramm",bg='red',fg='white',command=lambda:plot_int(folder_path.get(),selected_file))
    button4.place(relx=0.4, y=180, anchor='se')    

def compute_buttons():
    button5 = Button(text="Compute values from selected file",bg='green',fg='white',command=lambda:one_file(folder_path.get(),selected_file))
    button5.place(relx=0.65, y=180, anchor='se')
    button6 = Button(text="Analyse all files in Folder and get excel file",bg='green',fg='white',command=lambda :compute_data_from_all_files(folder_path.get(),save_var.get()))
    button6.place(x=770, y=180, anchor='se')

def sensorcallback(selection):
    
    global selected_sensor
    selected_sensor = selection
    print(selected_sensor)
    if (choice.get() == 'Innoviz Pro'):
        h_res_var.set(0.2)                                                                      
        v_res_var.set(0.45)
      
    elif(choice.get()== 'Cepton Vista P60'):
        h_res_var.set(0.25)                                                                      
        v_res_var.set(0.25)
        

def sensor_choice():
    list_of_sensors=['Innoviz Pro','Cepton Vista P60']
    global choice
    choice = StringVar()
    choice.set('Innoviz Pro') # default value
    filemenu = OptionMenu(root,choice, *list_of_sensors,command=sensorcallback)  
    filemenu.place(x=500, y=83, anchor='se')
    
    global h_res_var
    h_res_var = StringVar()          
    h_res_var.set(0.2)
    h_res_entry = Label(root, textvariable=h_res_var)   
    h_res_entry.place(x=670,y=90,anchor='w')                                         
    lbl40 = Label(root, text="H-Res of sensor")
    lbl40.place(x=559, y=90, anchor='w')
    global v_res_var
    v_res_var = StringVar()
    v_res_var.set(0.45)
    v_res_entry = Label(root, textvariable=v_res_var )
    v_res_entry.place(x=670,y=110,anchor='w')
    lbl50 = Label(root, text="V-Res of sensor")
    lbl50.place(x=559, y=110, anchor='w')

def show_buttons():
    folderselection()
    fileselection()
    sensor_choice()
    plot_dist_button()
    plot_int_button()
    entry_buttons_show()
    compute_buttons()
    
    buttontotest=Button(text='clickme',command=lambda:testfunction())
    buttontotest.pack()   
    
def testfunction():
    print(target_v_orig,target_h_orig,number_of_frames)
    #print(h_res , '   ',v_res)


def main():
    global root
    root= Tk()
    root.title('Test GUI')
    root.geometry("800x700")
    
    show_buttons()

    
    root.mainloop()
    
    
if __name__ == '__main__':
    main()